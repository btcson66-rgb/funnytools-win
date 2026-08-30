from __future__ import annotations

import csv
import io
from statistics import median

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .common import MAX_OCR_PAGES, open_pdf_reader, run_tesseract_png, validate_ocr_lang
from .pdf_render import render_pdf_page_png


def _cluster_rows(words: list[dict], tolerance: float = 5.0) -> list[list[dict]]:
    rows: list[list[dict]] = []
    for word in sorted(words, key=lambda w: (float(w["top"]), float(w["x0"]))):
        y = float(word["top"])
        target = None
        for row in rows:
            row_y = sum(float(w["top"]) for w in row) / len(row)
            if abs(row_y - y) <= tolerance:
                target = row
                break
        if target is None:
            target = []
            rows.append(target)
        target.append(word)
    for row in rows:
        row.sort(key=lambda w: float(w["x0"]))
    return rows


def _table_from_words(words: list[dict]) -> list[list[str]]:
    if len(words) < 4:
        return []
    rows = _cluster_rows(words)
    rows = [row for row in rows if len(row) >= 2]
    # Coordinate fallback is intentionally conservative: with only two text lines,
    # ordinary prose can look like aligned columns. Real two-row bordered tables are
    # normally caught by pdfplumber before this fallback is reached.
    if len(rows) < 3:
        return []

    # Detect recurring left-edge anchors. Real table columns usually start at
    # similar x positions across several rows; normal prose usually does not.
    heights = [float(w.get("bottom", w["top"])) - float(w["top"]) for w in words]
    positive_heights = [h for h in heights if h > 0]
    typical_height = median(positive_heights) if positive_heights else 12.0
    x_tolerance = max(12.0, typical_height * 0.65)

    clusters: list[dict] = []
    for row_index, row in enumerate(rows):
        for word in row:
            x = float(word["x0"])
            match = None
            for cluster in clusters:
                if abs(x - cluster["center"]) <= x_tolerance:
                    match = cluster
                    break
            if match is None:
                clusters.append({"center": x, "values": [x], "rows": {row_index}})
            else:
                match["values"].append(x)
                match["rows"].add(row_index)
                match["center"] = sum(match["values"]) / len(match["values"])

    min_occurrence = max(2, (len(rows) + 1) // 2)
    anchors = sorted(
        cluster["center"] for cluster in clusters if len(cluster["rows"]) >= min_occurrence
    )
    if len(anchors) < 2:
        return []

    # Merge near-duplicate anchors once more after averaging.
    merged: list[float] = []
    for x in anchors:
        if not merged or abs(x - merged[-1]) > x_tolerance:
            merged.append(x)
        else:
            merged[-1] = (merged[-1] + x) / 2
    anchors = merged
    if len(anchors) < 2:
        return []

    boundaries = [(a + b) / 2 for a, b in zip(anchors, anchors[1:])]
    table: list[list[str]] = []
    for row in rows:
        cells = [[] for _ in anchors]
        for word in row:
            center = (float(word["x0"]) + float(word["x1"])) / 2
            col = 0
            while col < len(boundaries) and center > boundaries[col]:
                col += 1
            cells[col].append(str(word["text"]))
        values = [" ".join(cell).strip() for cell in cells]
        if sum(bool(v) for v in values) >= 2:
            table.append(values)

    if len(table) < 3:
        return []
    return table


def _ocr_words(pdf_bytes: bytes, page_index: int, lang: str) -> list[dict]:
    png = render_pdf_page_png(pdf_bytes, page_index, dpi=250)
    tsv = run_tesseract_png(png, lang, output_format="tsv", psm=6).decode("utf-8", errors="replace")
    rows = csv.DictReader(io.StringIO(tsv), delimiter="\t")
    words: list[dict] = []
    for row in rows:
        text = (row.get("text") or "").strip()
        if not text:
            continue
        try:
            conf = float(row.get("conf") or -1)
        except ValueError:
            conf = -1
        if conf < 20:
            continue
        x = float(row.get("left") or 0)
        y = float(row.get("top") or 0)
        w = float(row.get("width") or 0)
        h = float(row.get("height") or 0)
        words.append({"text": text, "x0": x, "x1": x + w, "top": y, "bottom": y + h})
    return words


def extract_tables(
    pdf_bytes: bytes,
    filename: str = "tables.pdf",
    pages: list[int] | None = None,
    ocr_mode: str = "auto",
    ocr_lang: str = "eng",
) -> tuple[list[dict], dict]:
    reader = open_pdf_reader(pdf_bytes, filename)
    if ocr_mode not in {"auto", "force", "off"}:
        raise ValueError("ocr_mode must be auto, force, or off")
    validate_ocr_lang(ocr_lang)

    page_indices = pages if pages is not None else list(range(len(reader.pages)))
    page_indices = sorted(set(page_indices))
    for pno in page_indices:
        if pno < 0 or pno >= len(reader.pages):
            raise ValueError(f"Requested page {pno + 1} does not exist")

    result: list[dict] = []
    fallback_count = 0
    ocr_count = 0

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pno in page_indices:
            page = pdf.pages[pno]
            page_results = 0

            if ocr_mode != "force":
                try:
                    tables = page.extract_tables() or []
                except Exception:
                    tables = []
                for t_idx, rows in enumerate(tables, start=1):
                    cleaned = [["" if c is None else str(c).strip() for c in row] for row in rows]
                    cleaned = [row for row in cleaned if any(cell for cell in row)]
                    if len(cleaned) >= 2 and max((sum(bool(c) for c in row) for row in cleaned), default=0) >= 2:
                        result.append({"page": pno + 1, "table": t_idx, "method": "pdfplumber", "rows": cleaned})
                        page_results += 1

                if page_results == 0:
                    words = page.extract_words() or []
                    normalized = [
                        {
                            "text": str(w.get("text", "")),
                            "x0": float(w.get("x0", 0)),
                            "x1": float(w.get("x1", 0)),
                            "top": float(w.get("top", 0)),
                            "bottom": float(w.get("bottom", w.get("top", 0))),
                        }
                        for w in words
                    ]
                    fallback = _table_from_words(normalized)
                    if fallback:
                        result.append({"page": pno + 1, "table": 1, "method": "coordinate-fallback", "rows": fallback})
                        fallback_count += 1
                        page_results += 1

            needs_ocr = ocr_mode == "force" or (
                ocr_mode == "auto" and page_results == 0 and len((page.extract_text() or "").strip()) < 20
            )
            if needs_ocr:
                if ocr_count >= MAX_OCR_PAGES:
                    raise ValueError(f"OCR table extraction is limited to {MAX_OCR_PAGES} pages per job")
                ocr_words = _ocr_words(pdf_bytes, pno, ocr_lang)
                ocr_table = _table_from_words(ocr_words)
                if ocr_table:
                    result.append({"page": pno + 1, "table": 1, "method": "ocr-coordinate", "rows": ocr_table})
                    ocr_count += 1

    return result, {
        "pages_scanned": len(page_indices),
        "tables_found": len(result),
        "fallback_tables": fallback_count,
        "ocr_tables": ocr_count,
    }


def tables_to_xlsx(tables: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    if not tables:
        ws = wb.create_sheet("No tables")
        ws["A1"] = "No table detected"
        ws["A1"].font = Font(bold=True)
    else:
        for idx, item in enumerate(tables, start=1):
            page = int(item.get("page", 1))
            table_no = int(item.get("table", idx))
            title = f"P{page}_T{table_no}"[:31]
            if title in wb.sheetnames:
                title = f"{title[:27]}_{idx}"[:31]
            ws = wb.create_sheet(title)
            rows = item.get("rows") or []
            for r, row in enumerate(rows, start=1):
                if not isinstance(row, list):
                    raise ValueError("Each table row must be a list")
                for c, value in enumerate(row, start=1):
                    if isinstance(value, (dict, list)):
                        raise ValueError("Table cells must be scalar values")
                    ws.cell(r, c, "" if value is None else str(value))
            max_cols = max((len(r) for r in rows), default=1)
            for c in range(1, max_cols + 1):
                values = [str(ws.cell(r, c).value or "") for r in range(1, ws.max_row + 1)]
                width = min(50, max(10, max((len(v) for v in values), default=10) + 2))
                ws.column_dimensions[get_column_letter(c)].width = width

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def pdf_tables_to_xlsx(
    pdf_bytes: bytes,
    filename: str = "tables.pdf",
    pages: list[int] | None = None,
    ocr_mode: str = "auto",
    ocr_lang: str = "eng",
) -> tuple[bytes, list[dict], dict]:
    tables, stats = extract_tables(
        pdf_bytes,
        filename=filename,
        pages=pages,
        ocr_mode=ocr_mode,
        ocr_lang=ocr_lang,
    )
    return tables_to_xlsx(tables), tables, stats
